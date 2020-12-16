import openpyxl

def getrowcount(filename,sheetname):
    workbook=openpyxl.load_workbook(filename)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)

def getcolumncount(filename,sheetname):
    workbook=openpyxl.load_workbook(filename)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)

def readdata(filename,sheetname,rowno,colno):
    workbook=openpyxl.load_workbook(filename)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowno,column=colno).value

def writedata(filename,sheetname,rowno,colno,data):
    workbook=openpyxl.load_workbook(filename)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rowno,column=colno).value=data
    workbook.save(filename)